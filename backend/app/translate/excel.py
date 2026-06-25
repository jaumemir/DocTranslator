# translate/excel.py
import datetime
import logging
from typing import List, Dict, Any
from threading import Event
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from . import to_translate
from . import common


def start(trans: Dict[str, Any]) -> bool:
    """
    Excel file translation entry point
    :param trans: Translation configuration dictionary
    :return: Whether successful
    """
    translate_id = trans['id']
    start_time = datetime.datetime.now()

    # Load workbook
    try:
        wb = openpyxl.load_workbook(trans['file_path'])
    except Exception as e:
        logging.error(f"[Task {translate_id}] Unable to open Excel file: {e}")
        to_translate.error(translate_id, f"Unable to open Excel file: {str(e)}")
        return False

    # Extract all cells that need translation
    texts = []
    cell_map = []  # Record cell positions for write-back

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _extract_sheet_texts(ws, sheet_name, texts, cell_map)
    except Exception as e:
        logging.error(f"[Task {translate_id}] Text extraction failed: {e}")
        to_translate.error(translate_id, f"Text extraction failed: {str(e)}")
        return False

    if not texts:
        logging.info(f"[Task {translate_id}] No content to translate in Excel file")
        wb.save(trans['target_file'])
        to_translate.complete(trans, 0, "0s")
        return True

    logging.info(f"[Task {translate_id}] Extracted {len(texts)} cells for translation")

    # [Key modification] Use thread pool for batch translation
    event = Event()
    success = to_translate.translate_batch(trans, texts, event)
    if not success:
        return False

    # Write back translation results
    try:
        text_count = _apply_translation(wb, texts, cell_map, trans.get('type', ''))
        wb.save(trans['target_file'])
    except Exception as e:
        logging.error(f"[Task {translate_id}] Failed to save file: {e}")
        to_translate.error(translate_id, f"Failed to save file: {str(e)}")
        return False

    end_time = datetime.datetime.now()
    spend_time = common.display_spend(start_time, end_time)
    to_translate.complete(trans, text_count, spend_time)
    return True


def _extract_sheet_texts(ws: Worksheet, sheet_name: str, texts: List[Dict], cell_map: List[Dict]):
    """
    Extract text from worksheet
    :param ws: Worksheet object
    :param sheet_name: Worksheet name
    :param texts: Text list (output)
    :param cell_map: Cell position mapping (output)
    """
    # Get merged cell ranges to avoid duplicate translation
    merged_cells = set()
    for merged_range in ws.merged_cells.ranges:
        # Only keep the first cell of merged area, skip others
        first_cell = True
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if first_cell:
                    first_cell = False
                else:
                    merged_cells.add((row, col))

    # Iterate through all cells
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            # Skip non-first cells of merged cells
            if (row_idx, col_idx) in merged_cells:
                continue

            value = cell.value
            if _should_translate(value):
                text_item = {
                    'text': str(value),
                    'original': str(value),
                    'complete': False,
                    'count': 0
                }
                texts.append(text_item)
                cell_map.append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'col': col_idx,
                    'text_index': len(texts) - 1
                })


def _should_translate(value) -> bool:
    """Determine if cell value needs translation"""
    if value is None:
        return False
    if isinstance(value, (int, float, complex)):
        return False
    if isinstance(value, datetime.datetime):
        return False
    if isinstance(value, datetime.time):
        return False

    text = str(value).strip()
    if not text:
        return False
    if common.is_all_punc(text):
        return False

    return True



def _apply_translation(wb, texts: List[Dict], cell_map: List[Dict], trans_type: str) -> int:
    """
    Apply translation results to workbook
    :return: Translation word count statistics
    """
    text_count = 0
    keep_both = 'both' in trans_type

    for mapping in cell_map:
        sheet_name = mapping['sheet']
        row = mapping['row']
        col = mapping['col']
        text_index = mapping['text_index']

        if text_index >= len(texts):
            continue

        text_item = texts[text_index]
        text_count += text_item.get('count', 0)

        ws = wb[sheet_name]
        cell = ws.cell(row=row, column=col)

        original = text_item.get('original', '')
        translated = text_item.get('text', original)

        if keep_both:
            # Keep both original and translation, separated by newline
            cell.value = f"{original}\n{translated}"
        else:
            # Keep only translation
            cell.value = translated

    return text_count
